from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse  # Import JsonResponse for returning JSON responses
from django.contrib.auth import logout, authenticate, login  # Import necessary functions for user authentication
from django.contrib.auth.models import User  # Import User model for registration
from django.contrib.auth.tokens import default_token_generator  # Import token generator for password reset
from django.utils.http import urlsafe_base64_encode  # Import for encoding user ID
from django.utils.encoding import force_bytes  # Import for encoding user ID
from django.core.mail import send_mail  # Import for sending emails
from django.views.decorators.csrf import csrf_exempt  # Import csrf_exempt decorator
from django.contrib import messages  # Import messages framework
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .forms import UserRegistrationForm, WeeklyDayOffForm, LeaveRequestForm  # Import the registration, weekly day off, and leave request forms
from .models import LeaveRequest, AttendanceRecord, WeeklyDayOff  # Import the LeaveRequest, AttendanceRecord, and WeeklyDayOff models
import datetime  # Import the datetime module

@login_required
def office_head_dashboard(request):
    employees = User.objects.filter(is_staff=False, is_superuser=False)

    leave_requests_list = LeaveRequest.objects.all().order_by('-applied_at')
    attendance_records_list = AttendanceRecord.objects.select_related('user').order_by('-date')
    weekly_day_offs_list = WeeklyDayOff.objects.select_related('user').order_by('user__username')

    # Pagination for leave requests
    leave_page = request.GET.get('leave_page', 1)
    leave_paginator = Paginator(leave_requests_list, 10)
    try:
        leave_requests = leave_paginator.page(leave_page)
    except PageNotAnInteger:
        leave_requests = leave_paginator.page(1)
    except EmptyPage:
        leave_requests = leave_paginator.page(leave_paginator.num_pages)

    # Pagination for attendance records
    attendance_page = request.GET.get('attendance_page', 1)
    attendance_paginator = Paginator(attendance_records_list, 10)
    try:
        attendance_records = attendance_paginator.page(attendance_page)
    except PageNotAnInteger:
        attendance_records = attendance_paginator.page(1)
    except EmptyPage:
        attendance_records = attendance_paginator.page(attendance_paginator.num_pages)

    # Pagination for weekly day offs
    weekly_day_off_page = request.GET.get('weekly_day_off_page', 1)
    weekly_day_off_paginator = Paginator(weekly_day_offs_list, 10)
    try:
        weekly_day_offs = weekly_day_off_paginator.page(weekly_day_off_page)
    except PageNotAnInteger:
        weekly_day_offs = weekly_day_off_paginator.page(1)
    except EmptyPage:
        weekly_day_offs = weekly_day_off_paginator.page(weekly_day_off_paginator.num_pages)

    return render(request, 'attendance/office_head_dashboard.html', {
        'employees': employees,
        'leave_requests': leave_requests,
        'attendance_records': attendance_records,
        'weekly_day_offs': weekly_day_offs,
        'is_office_head': True,
    })

@login_required
def dashboard(request):
    attendance_records_list = AttendanceRecord.objects.filter(user=request.user).order_by('-date')
    leave_requests_list = LeaveRequest.objects.filter(user=request.user).order_by('-applied_at')

    # Pagination for attendance records
    attendance_page = request.GET.get('attendance_page', 1)
    attendance_paginator = Paginator(attendance_records_list, 10)
    try:
        attendance_records = attendance_paginator.page(attendance_page)
    except PageNotAnInteger:
        attendance_records = attendance_paginator.page(1)
    except EmptyPage:
        attendance_records = attendance_paginator.page(attendance_paginator.num_pages)

    # Pagination for leave requests
    leave_page = request.GET.get('leave_page', 1)
    leave_paginator = Paginator(leave_requests_list, 10)
    try:
        leave_requests = leave_paginator.page(leave_page)
    except PageNotAnInteger:
        leave_requests = leave_paginator.page(1)
    except EmptyPage:
        leave_requests = leave_paginator.page(leave_paginator.num_pages)

    # Calculate work_hours for each attendance record
    for record in attendance_records:
        if record.check_in_time and record.check_out_time:
            check_in_dt = datetime.datetime.combine(record.date, record.check_in_time)
            check_out_dt = datetime.datetime.combine(record.date, record.check_out_time)
            duration = check_out_dt - check_in_dt
            # Convert duration to hours and minutes string
            hours, remainder = divmod(duration.seconds, 3600)
            minutes, _ = divmod(remainder, 60)
            record.work_hours = f"{hours}h {minutes}m"
        else:
            record.work_hours = "N/A"

    try:
        weekly_day_off = WeeklyDayOff.objects.get(user=request.user)
    except WeeklyDayOff.DoesNotExist:
        weekly_day_off = None

    return render(request, 'attendance/dashboard.html', {
        'attendance_records': attendance_records,
        'weekly_day_off': weekly_day_off,
        'leave_requests': leave_requests,
        'is_office_head': False,
    })

@login_required
@require_POST
def approve_leave(request, leave_id):
    leave_request = get_object_or_404(LeaveRequest, id=leave_id)
    leave_request.status = 'Approved'
    leave_request.save()
    return redirect('attendance:office_head_dashboard')

@login_required
@require_POST
def reject_leave(request, leave_id):
    leave_request = get_object_or_404(LeaveRequest, id=leave_id)
    leave_request.status = 'Rejected'
    leave_request.save()
    return redirect('attendance:office_head_dashboard')

@login_required
def mark_attendance(request):
    today = datetime.date.today()
    record, created = AttendanceRecord.objects.get_or_create(user=request.user, date=today)
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'check_in':
            if not record.check_in_time:
                record.check_in_time = datetime.datetime.now().time()
                record.save()
            return redirect('attendance:dashboard')
        elif action == 'check_out':
            if record.check_in_time and not record.check_out_time:
                record.check_out_time = datetime.datetime.now().time()
                record.save()
            return redirect('attendance:dashboard')
    return render(request, 'attendance/mark_attendance.html', {'record': record})

@login_required
def user_logout(request):
    logout(request)
    return redirect('attendance:login')

@login_required
def leave_request_create(request):
    if request.method == 'POST':
        form = LeaveRequestForm(request.POST)
        if form.is_valid():
            leave_request = form.save(commit=False)
            leave_request.user = request.user
            leave_request.save()
            return redirect('attendance:leave_request_list')
    else:
        form = LeaveRequestForm()
    return render(request, 'attendance/leave_request_form.html', {'form': form})

@login_required
def weekly_day_off(request):
    if request.method == 'POST':
        form = WeeklyDayOffForm(request.POST)
        if form.is_valid():
            day_off = form.cleaned_data['day_off']
            weekly_day_off_obj, created = WeeklyDayOff.objects.update_or_create(
                user=request.user,
                defaults={'day_off': day_off}
            )
            return redirect('attendance:dashboard')
    else:
        try:
            weekly_day_off_obj = WeeklyDayOff.objects.get(user=request.user)
            form = WeeklyDayOffForm(initial={'day_off': weekly_day_off_obj.day_off})
        except WeeklyDayOff.DoesNotExist:
            form = WeeklyDayOffForm()
    return render(request, 'attendance/weekly_day_off.html', {'form': form})

@csrf_exempt
def user_login(request):
    if request.user.is_authenticated:
        return redirect('attendance:dashboard')
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, 'Login successful!')
            return redirect('attendance:dashboard')
        else:
            messages.error(request, 'Invalid username or password')
            return render(request, 'attendance/login.html')
    return render(request, 'attendance/login.html')

def register(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            return redirect('attendance:login')
    else:
        form = UserRegistrationForm()
    return render(request, 'attendance/register.html', {'form': form})

from django.contrib import messages

def office_head_login(request):
    success_message = request.GET.get('success_message')
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('attendance:office_head_dashboard')
        else:
            return render(request, 'attendance/office_head_login.html', {'error': 'Invalid username or password'})
    return render(request, 'attendance/office_head_login.html', {'success_message': success_message})

def office_head_register(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            return redirect('attendance:office_head_login')
    else:
        form = UserRegistrationForm()
    return render(request, 'attendance/office_head_register.html', {'form': form})

@login_required
def leave_request_list(request):
    leave_requests = LeaveRequest.objects.filter(user=request.user).order_by('-applied_at')
    return render(request, 'attendance/leave_request_list.html', {'leave_requests': leave_requests})

@login_required
def leave_request_cancel(request, leave_id):
    leave_request = get_object_or_404(LeaveRequest, id=leave_id, user=request.user)
    if leave_request.status == 'Pending':
        leave_request.status = 'Cancelled'
        leave_request.save()
    return redirect('attendance:leave_request_list')

from django.contrib.auth.forms import PasswordResetForm
from django.core.mail import BadHeaderError
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.contrib.auth.tokens import default_token_generator
from django.shortcuts import redirect

def password_reset(request):
    if request.method == "POST":
        password_reset_form = PasswordResetForm(request.POST)
        if password_reset_form.is_valid():
            data = password_reset_form.cleaned_data['email']
            associated_users = User.objects.filter(email=data)
            if associated_users.exists():
                for user in associated_users:
                    subject = "Password Reset Requested"
                    email_template_name = "attendance/password_reset_email.txt"
                    c = {
                        "email": user.email,
                        'domain': request.get_host(),
                        'site_name': 'Attendance System',
                        "uid": urlsafe_base64_encode(force_bytes(user.pk)),
                        "user": user,
                        'token': default_token_generator.make_token(user),
                        'protocol': 'http',
                    }
                    email = render_to_string(email_template_name, c)
                    try:
                        user.email_user(subject, email)
                    except BadHeaderError:
                        return HttpResponse('Invalid header found.')
                return redirect("/password-reset/done/")
    password_reset_form = PasswordResetForm()
    return render(request=request, template_name="attendance/password_reset.html", context={"password_reset_form":password_reset_form})

import csv
from django.http import HttpResponse

@login_required
def export_attendance_records(request):
    # Export attendance records as CSV for the logged-in user
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="attendance_records.csv"'

    writer = csv.writer(response)
    writer.writerow(['Date', 'Check In Time', 'Check Out Time', 'Work Hours'])

    attendance_records = AttendanceRecord.objects.filter(user=request.user).order_by('-date')
    for record in attendance_records:
        if record.check_in_time and record.check_out_time:
            check_in_dt = datetime.datetime.combine(record.date, record.check_in_time)
            check_out_dt = datetime.datetime.combine(record.date, record.check_out_time)
            duration = check_out_dt - check_in_dt
            hours, remainder = divmod(duration.seconds, 3600)
            minutes, _ = divmod(remainder, 60)
            work_hours = f"{hours}h {minutes}m"
        else:
            work_hours = "N/A"
        writer.writerow([record.date, record.check_in_time, record.check_out_time, work_hours])

    return response

import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse

@login_required
def export_office_head_data(request):
    # Create an in-memory output file for the new workbook.
    output = io.BytesIO()
    wb = Workbook()

    # Attendance Records Sheet
    ws1 = wb.active
    ws1.title = "Attendance Records"
    attendance_headers = ['Employee', 'Date', 'Check In Time', 'Check Out Time', 'Work Hours']
    ws1.append(attendance_headers)
    for col_num, header in enumerate(attendance_headers, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    attendance_records = AttendanceRecord.objects.select_related('user').order_by('user__username', '-date')
    for record in attendance_records:
        if record.check_in_time and record.check_out_time:
            check_in_dt = datetime.datetime.combine(record.date, record.check_in_time)
            check_out_dt = datetime.datetime.combine(record.date, record.check_out_time)
            duration = check_out_dt - check_in_dt
            hours, remainder = divmod(duration.seconds, 3600)
            minutes, _ = divmod(remainder, 60)
            work_hours = f"{hours}h {minutes}m"
        else:
            work_hours = "N/A"
        ws1.append([record.user.username, record.date.strftime('%Y-%m-%d'), str(record.check_in_time), str(record.check_out_time), work_hours])

    # Weekly Day Off Sheet
    ws2 = wb.create_sheet(title="Weekly Day Off")
    day_off_headers = ['Employee', 'Weekly Day Off']
    ws2.append(day_off_headers)
    for col_num, header in enumerate(day_off_headers, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    weekly_day_offs = WeeklyDayOff.objects.select_related('user').order_by('user__username')
    for wdo in weekly_day_offs:
        ws2.append([wdo.user.username, wdo.day_off])

    # Leave Requests Sheet
    ws3 = wb.create_sheet(title="Leave Requests")
    leave_headers = ['Employee', 'Leave Type', 'Start Date', 'End Date', 'Status', 'Applied At']
    ws3.append(leave_headers)
    for col_num, header in enumerate(leave_headers, 1):
        cell = ws3.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    leave_requests = LeaveRequest.objects.select_related('user').order_by('user__username', '-applied_at')
    for lr in leave_requests:
        ws3.append([
            lr.user.username,
            lr.leave_type,
            lr.start_date.strftime('%Y-%m-%d'),
            lr.end_date.strftime('%Y-%m-%d'),
            lr.status,
            lr.applied_at.strftime('%Y-%m-%d %H:%M:%S')
        ])

    # Adjust column widths for all sheets
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    wb.save(output)
    output.seek(0)

    filename = "office_head_data_export.xlsx"
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    return response
